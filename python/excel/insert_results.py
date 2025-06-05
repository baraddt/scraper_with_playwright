import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


def append_rows_to_excel(
    filepath, rows, url, sold=None, rating=None, rating_count=None
):

    is_pivot_format = False
    if rows and isinstance(rows[0], dict):
        first_row_keys = rows[0].keys()
        if (
            "Masa Aktif" in first_row_keys
            and ("Kuota" in first_row_keys or "FUP" in first_row_keys)
            and "Harga" in first_row_keys
        ):
            is_pivot_format = True

    existing_data_list = []
    if os.path.exists(filepath):
        df_existing_raw = pd.read_excel(filepath, header=None)
        existing_data_list = df_existing_raw.values.tolist()

    new_block_data = []

    # --- [1] Tulis URL dan Info meta ---
    new_block_data.append([url])
    new_block_data.append(
        ["Sold", sold, "Rating", rating, "Rating Count", rating_count]
    )
    new_block_data.append([])

    # --- [2] Pivot-style ---
    if is_pivot_format:
        df_temp_rows = []
        for row in rows:
            masa_aktif = row.get("Masa Aktif", "")
            kuota_key = "Kuota" if "Kuota" in row else "FUP"
            kuota = row.get(kuota_key, "")
            harga = row.get("Harga", "")

            if masa_aktif and kuota and harga:
                df_temp_rows.append(
                    {"Masa Aktif": masa_aktif, "Kuota": kuota, "Harga": harga}
                )

        if not df_temp_rows:
            print(
                f"No valid data found for pivot table format from URL: {url}. Skipping."
            )
            return

        df_new_raw = pd.DataFrame(df_temp_rows)

        df_new_raw["Harga"] = (
            df_new_raw["Harga"]
            .astype(str)
            .str.replace("Rp", "", regex=False)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_new_raw["Harga"] = pd.to_numeric(df_new_raw["Harga"], errors="coerce")
        df_new_raw.dropna(subset=["Harga"], inplace=True)

        if df_new_raw.empty:
            print(
                f"No valid numeric prices found for pivot table from URL: {url}. Skipping."
            )
            return

        df_pivot = df_new_raw.pivot_table(
            index="Masa Aktif", columns="Kuota", values="Harga", aggfunc="first"
        )
        df_pivot = df_pivot.reset_index()

        new_block_data.append(df_pivot.columns.tolist())
        for _, row_data in df_pivot.iterrows():
            new_block_data.append(row_data.tolist())

    # --- [3] Normal-style ---
    else:
        cols_to_use = []
        if any("Nama Paket" in r for r in rows):
            cols_to_use = ["Nama Produk", "Harga"]
        elif any("Masa Aktif - Kuota" in r for r in rows):
            cols_to_use = ["Nama Produk", "Harga"]
        else:
            if rows:
                cols_to_use = list(rows[0].keys())

        new_block_data.append(cols_to_use)
        for row in rows:
            formatted_row = []
            for col in cols_to_use:
                if col == "Nama Produk":
                    if "Nama Paket" in row:
                        formatted_row.append(row.get("Nama Paket", ""))
                    elif "Masa Aktif - Kuota" in row:
                        formatted_row.append(row.get("Masa Aktif - Kuota", ""))
                    else:
                        formatted_row.append(row.get(col, ""))
                elif col == "Harga":
                    formatted_row.append(row.get("Harga", ""))
                else:
                    formatted_row.append(row.get(col, ""))
            new_block_data.append(formatted_row)

    # --- [4] Tambahkan pemisah dan gabungkan ke data lama ---
    if existing_data_list:
        max_width_existing = (
            max(len(row) for row in existing_data_list) if existing_data_list else 0
        )
        max_width_new_block = (
            max(len(row) for row in new_block_data) if new_block_data else 0
        )
        separator_width = max(max_width_existing, max_width_new_block, 1)

        existing_data_list.append([None] * separator_width)
        existing_data_list.append([None] * separator_width)

    existing_data_list.extend(new_block_data)
    df_final = pd.DataFrame(existing_data_list)
    df_final.to_excel(filepath, index=False, header=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # Tentukan baris dan kolom untuk seluruh data (baik lama dan baru)
    total_rows = len(existing_data_list)
    total_cols = max(len(row) for row in existing_data_list)

    # Define thick border style
    thick_border = Border(
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000"),
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000"),
    )

    # Apply border to all blocks (using dynamic row detection)
    row_start = 1  # Baris pertama
    for row in range(1, total_rows + 1):
        if existing_data_list[row - 1] == [None] * total_cols:  # Deteksi baris kosong
            row_end = row - 1  # Selesai dengan blok sebelumnya
            # Terapkan border di blok produk sebelumnya (antara row_start dan row_end)
            for col in range(1, total_cols + 1):
                # Terapkan border di luar blok
                cell = ws.cell(row=row_start, column=col)
                cell.border = Border(top=thick_border.top if col == 1 else None)

                cell = ws.cell(row=row_end, column=col)
                cell.border = Border(
                    bottom=thick_border.bottom if col == total_cols else None
                )
            row_start = (
                row + 1
            )  # Tentukan baris berikutnya sebagai start row untuk blok selanjutnya

    # Terapkan border pada blok terakhir
    row_end = total_rows
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row_start, column=col)
        cell.border = Border(top=thick_border.top if col == 1 else None)

        cell = ws.cell(row=row_end, column=col)
        cell.border = Border(bottom=thick_border.bottom if col == total_cols else None)

    # Save the workbook with borders applied
    wb.save(filepath)
